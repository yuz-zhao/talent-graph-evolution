# -*- coding: utf-8 -*-
"""验证「100 条真实岗位 JD 测试数据」交付包（4 个 xlsx）的完整性与无样式约束。

用法：node scripts/run-python-script.mjs crawler/scripts/validate_jd_test_kit.py
退出码：0 = 全部通过；1 = 存在失败项。
"""
import os
import sys

from openpyxl import load_workbook

OUT_DIR = r"E:\TalentGraph\材料\100条真实岗位JD测试数据"

CHECKS = {
    "100条JD原始数据.xlsx": 100,
    "100条JD人工标注.xlsx": 100,
    "JD测试用例.xlsx": 15,
    "JD来源清单.xlsx": 8,
}


def main():
    failures = []
    for fname, min_rows in CHECKS.items():
        path = os.path.join(OUT_DIR, fname)
        if not os.path.exists(path):
            failures.append(f"缺失文件 {fname}")
            print(f"[FAIL] 缺失 {fname}")
            continue
        wb = load_workbook(path)
        ws = wb.active
        rows = ws.max_row - 1
        styled = sum(1 for row in ws.iter_rows() for c in row if c.has_style)
        merged = len(ws.merged_cells.ranges)
        ok = rows >= min_rows and styled == 0 and merged == 0
        print(f"[{'PASS' if ok else 'FAIL'}] {fname}: rows={rows} styled={styled} merged={merged}")
        if rows < min_rows:
            failures.append(f"{fname} 行数不足 {rows}<{min_rows}")
        if styled:
            failures.append(f"{fname} 存在 {styled} 个带样式单元格")
        if merged:
            failures.append(f"{fname} 存在 {merged} 个合并单元格")

    # 原始数据关键列非空（文件缺失时已在主循环记录，此处跳过）
    path = os.path.join(OUT_DIR, "100条JD原始数据.xlsx")
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[1]]
        for key in ("岗位名称", "公司名称", "来源URL", "标准岗位名称"):
            if key not in header:
                failures.append(f"原始数据缺列 {key}")
                print(f"[FAIL] 原始数据缺列 {key}")
                continue
            idx = header.index(key)
            n = sum(1 for row in ws.iter_rows(min_row=2, values_only=True)
                    if row[idx] and str(row[idx]).strip())
            ok = n >= 100
            print(f"[{'PASS' if ok else 'FAIL'}] 原始数据.{key} 非空: {n}/100")
            if not ok:
                failures.append(f"{key} 非空不足 {n}")

        # 来源多样性（≥6 个来源）
        if "来源名称" in header:
            idx = header.index("来源名称")
            sources = {row[idx] for row in ws.iter_rows(min_row=2, values_only=True) if row[idx]}
            ok = len(sources) >= 6
            print(f"[{'PASS' if ok else 'FAIL'}] 来源多样性: {len(sources)} 个来源")
            if not ok:
                failures.append(f"来源数 {len(sources)} < 6")
        else:
            failures.append("原始数据缺列 来源名称")
            print("[FAIL] 原始数据缺列 来源名称")

    # 人工标注 skill 覆盖（文件缺失时已在主循环记录，此处跳过）
    path2 = os.path.join(OUT_DIR, "100条JD人工标注.xlsx")
    if os.path.exists(path2):
        wb2 = load_workbook(path2)
        ws2 = wb2.active
        h2 = [c.value for c in ws2[1]]
        if "必备技能(required)" in h2:
            i2 = h2.index("必备技能(required)")
            n_req = sum(1 for row in ws2.iter_rows(min_row=2, values_only=True)
                        if row[i2] and str(row[i2]).strip())
            print(f"[INFO] 人工标注 required 非空: {n_req}/100")

    print()
    if failures:
        print("RESULT: FAIL")
        for f in failures:
            print("  -", f)
        sys.exit(1)
    print("RESULT: ALL PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
